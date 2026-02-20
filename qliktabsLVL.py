from __future__ import annotations

import logging
import time
import platform
import ctypes
import urllib.parse
from ctypes import wintypes
import os
import glob
import json
from pathlib import Path
from datetime import datetime, timedelta
import os as _os
import re

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

LOG = logging.getLogger(__name__)

tiempo=30
corto_tiempo=2

def setup_driver() -> webdriver.Chrome:
	opts = Options()
	opts.add_argument("--no-sandbox")
	opts.add_argument("--disable-dev-shm-usage")
	service = Service(ChromeDriverManager().install())
	driver = webdriver.Chrome(service=service, options=opts)
	driver.maximize_window()
	return driver


def type_like_keyboard(driver: webdriver.Chrome, text: str, delay: float = 0.08, click_first: bool = True) -> bool:
	try:
		if click_first:
			try:
				active = driver.switch_to.active_element
				if not active or active.tag_name.lower() == 'html':
					try:
						driver.find_element(By.TAG_NAME, 'body').click()
					except Exception:
						pass
					time.sleep(2)
			except Exception:
				time.sleep(2)

		active = driver.switch_to.active_element
		if not active:
			return False

		sent = 0
		for ch in text:
			try:
				active.send_keys(ch)
				sent += 1
				time.sleep(delay)
			except Exception:
				continue
		LOG.info("Typist: enviados %d/%d caracteres", sent, len(text))
		return sent > 0
	except Exception:
		LOG.exception("Error en type_like_keyboard")
		return False


def _send_text_windows(text: str, delay: float = 0.08) -> None:
	if platform.system() != "Windows":
		return
	user32 = ctypes.WinDLL('user32', use_last_error=True)
	VkKeyScanW = user32.VkKeyScanW
	keybd_event = user32.keybd_event
	VK_SHIFT = 0x10
	KEYEVENTF_KEYUP = 0x0002
	for ch in text:
		vks = VkKeyScanW(ord(ch))
		if vks == -1:
			continue
		vk = vks & 0xFF
		shift_state = (vks >> 8) & 0xFF
		if shift_state & 1:
			keybd_event(VK_SHIFT, 0, 0, 0)
		keybd_event(vk, 0, 0, 0)
		time.sleep(0.01)
		keybd_event(vk, 0, KEYEVENTF_KEYUP, 0)
		if shift_state & 1:
			keybd_event(VK_SHIFT, 0, KEYEVENTF_KEYUP, 0)
		time.sleep(delay)


def send_text_via_system(driver: webdriver.Chrome, text: str, delay: float = 0.08) -> bool:
	try:
		try:
			driver.execute_script("window.focus();")
		except Exception:
			pass
		time.sleep(0.2)
		if platform.system() == "Windows":
			_send_text_windows(text, delay=delay)
			LOG.info("Envío de texto por sistema completado (Windows)")
			return True
		return False
	except Exception:
		LOG.exception("Fallo en send_text_via_system")
		return False


def send_keys_via_pywinauto(keys: str, url_substring: str | None = None) -> bool:
	try:
		from pywinauto import Desktop, keyboard
	except Exception:
		LOG.debug('pywinauto no instalado')
		return False
	try:
		desktop = Desktop(backend="uia")
		windows = desktop.windows()
		target = None
		us = url_substring.lower() if url_substring else None
		for w in windows:
			try:
				title = (w.window_text() or '').lower()
				if not title:
					continue
				if us and us in title:
					target = w
					break
				if not us and ('chrome' in title or 'qlik' in title):
					target = w
					break
			except Exception:
				continue
		if not target:
			for w in windows:
				try:
					title = (w.window_text() or '').lower()
					if 'qlik' in title or 'chrome' in title:
						target = w
						break
				except Exception:
					continue
		if not target:
			return False
		target.set_focus()
		time.sleep(0.12)
		keyboard.send_keys(keys)
		time.sleep(0.12)
		LOG.info('send_keys_via_pywinauto: keys enviados')
		return True
	except Exception:
		LOG.debug('send_keys_via_pywinauto: fallo', exc_info=True)
		return False


def _send_enter_windows() -> bool:
	if platform.system() != 'Windows':
		return False
	try:
		user32 = ctypes.WinDLL('user32', use_last_error=True)
		keybd_event = user32.keybd_event
		KEYEVENTF_KEYUP = 0x0002
		VK_RETURN = 0x0D
		keybd_event(VK_RETURN, 0, 0, 0)
		time.sleep(1)
		keybd_event(VK_RETURN, 0, KEYEVENTF_KEYUP, 0)
		time.sleep(1)
		LOG.info('_send_enter_windows: Enter enviado')
		return True
	except Exception:
		LOG.debug('_send_enter_windows: fallo', exc_info=True)
		return False


def focus_on_selector(driver: webdriver.Chrome, selector: str, timeout: float = 3.0) -> bool:
	try:
		end = time.time() + float(timeout)
		while time.time() < end:
			try:
				el = driver.find_element(By.CSS_SELECTOR, selector)
			except Exception:
				el = None
			if not el:
				time.sleep(1)
				continue

			try:
				driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'nearest'});", el)
			except Exception:
				pass

			try:
				el.click()
			except Exception:
				try:
					ActionChains(driver).move_to_element(el).click().perform()
				except Exception:
					try:
						driver.execute_script("arguments[0].focus();", el)
					except Exception:
						pass

			try:
				el.send_keys("")
			except Exception:
				pass

			try:
				active = driver.switch_to.active_element
				if active is not None:
					return True
				else:
					return True
			except Exception:
				return True

		return False
	except Exception:
		LOG.exception("focus_on_selector: excepción inesperada")
		return False


def hover_on_selector(driver: webdriver.Chrome, selector: str, timeout: float = 5.0) -> bool:
	try:
		end = time.time() + float(timeout)
		while time.time() < end:
			try:
				el = driver.find_element(By.CSS_SELECTOR, selector)
			except Exception:
				el = None
			if not el:
				time.sleep(0.5)
				continue

			try:
				driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'nearest'});", el)
			except Exception:
				pass

			try:
				ActionChains(driver).move_to_element(el).perform()
			except Exception:
				LOG.debug("hover_on_selector: ActionChains falló", exc_info=True)

			try:
				js = (
					"var e = new MouseEvent('mouseover', {bubbles:true, cancelable:true});"
					"arguments[0].dispatchEvent(e);"
					"var e2 = new MouseEvent('mouseenter', {bubbles:true, cancelable:true});"
					"arguments[0].dispatchEvent(e2);"
				)
				driver.execute_script(js, el)
			except Exception:
				LOG.debug("hover_on_selector: dispatch JS falló", exc_info=True)

			time.sleep(0.2)
			return True

		return False
	except Exception:
		LOG.exception("hover_on_selector: excepción inesperada")
		return False


def hover_on_xpath(driver: webdriver.Chrome, xpath: str, timeout: float = 5.0) -> bool:
	try:
		end = time.time() + float(timeout)
		while time.time() < end:
			try:
				el = driver.find_element(By.XPATH, xpath)
			except Exception:
				el = None
			if not el:
				time.sleep(1)
				continue

			try:
				driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'nearest'});", el)
			except Exception:
				pass

			try:
				ActionChains(driver).move_to_element(el).perform()
				LOG.info("hover_on_xpath: move_to_element realizado para %s", xpath)
			except Exception:
				LOG.debug("hover_on_xpath: ActionChains falló", exc_info=True)

			try:
				js = (
					"var e = new MouseEvent('mouseover', {bubbles:true, cancelable:true});"
					"arguments[0].dispatchEvent(e);"
					"var e2 = new MouseEvent('mouseenter', {bubbles:true, cancelable:true});"
					"arguments[0].dispatchEvent(e2);"
				)
				driver.execute_script(js, el)
				LOG.info("hover_on_xpath: eventos mouseover/mouseenter despachados para %s", xpath)
			except Exception:
				LOG.debug("hover_on_xpath: dispatch JS falló", exc_info=True)

			time.sleep(corto_tiempo)
			return True

		return False
	except Exception:
		LOG.exception("hover_on_xpath: excepción inesperada")
		return False


def click_button_by_selector(driver: webdriver.Chrome, selector: str, timeout: float = 5.0) -> bool:
	try:
		end = time.time() + float(timeout)
		while time.time() < end:
			try:
				btn = driver.find_element(By.CSS_SELECTOR, selector)
			except Exception:
				btn = None
			if not btn:
				time.sleep(0.25)
				continue

			try:
				driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'nearest'});", btn)
			except Exception:
				pass

			try:
				btn.click()
				LOG.info("click_button_by_selector: click directo en %s", selector)
				return True
			except Exception:
				try:
					ActionChains(driver).move_to_element(btn).click().perform()
					LOG.info("click_button_by_selector: click via ActionChains en %s", selector)
					return True
				except Exception:
					try:
						driver.execute_script("arguments[0].click();", btn)
						LOG.info("click_button_by_selector: click via JS en %s", selector)
						return True
					except Exception:
						LOG.debug("click_button_by_selector: intento de click falló para %s", selector, exc_info=True)
			time.sleep(0.2)
		return False
	except Exception:
		LOG.exception("click_button_by_selector: excepción inesperada")
		return False


def click_button_by_xpath(driver: webdriver.Chrome, xpath: str, timeout: float = 5.0) -> bool:
	try:
		end = time.time() + float(timeout)
		while time.time() < end:
			try:
				btn = driver.find_element(By.XPATH, xpath)
			except Exception:
				btn = None
			if not btn:
				time.sleep(0.25)
				continue

			try:
				driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'nearest'});", btn)
			except Exception:
				pass

			try:
				btn.click()
				LOG.info("click_button_by_xpath: click directo en %s", xpath)
				return True
			except Exception:
				try:
					ActionChains(driver).move_to_element(btn).click().perform()
					LOG.info("click_button_by_xpath: click via ActionChains en %s", xpath)
					return True
				except Exception:
					try:
						driver.execute_script("arguments[0].click();", btn)
						LOG.info("click_button_by_xpath: click via JS en %s", xpath)
						return True
					except Exception:
						LOG.debug("click_button_by_xpath: intento de click falló para %s", xpath, exc_info=True)
			time.sleep(0.2)
		return False
	except Exception:
		LOG.exception("click_button_by_xpath: excepción inesperada")
		return False


def click_export_url(driver: webdriver.Chrome, selector: str = 'a.export-url', timeout: float = 10.0) -> bool:
	try:
		end = time.time() + float(timeout)
		while time.time() < end:
			try:
				a = driver.find_element(By.CSS_SELECTOR, selector)
			except Exception:
				a = None
			if not a:
				time.sleep(0.5)
				continue

			try:
				href = a.get_attribute('href') or a.get_attribute('ng-href')
			except Exception:
				href = None

			try:
				driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", a)
			except Exception:
				pass

			try:
				driver.execute_script("arguments[0].click();", a)
				LOG.info("click_export_url: clicado enlace %s", selector)
				return True
			except Exception:
				try:
					a.click()
					LOG.info("click_export_url: clicado elemento via click() %s", selector)
					return True
				except Exception:
					LOG.debug("click_export_url: intento de click falló para %s", selector, exc_info=True)

			time.sleep(0.2)
		return False
	except Exception:
		LOG.exception("click_export_url: excepción inesperada")
		return False


def click_export_link_with_fallback(driver: webdriver.Chrome, timeout: float = 6.0) -> bool:
	try:
		if click_export_url(driver, selector='a.export-url', timeout=2.0):
			return True

		end = time.time() + float(timeout)
		xpaths = [
			"//a[contains(translate(@href,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'.xlsx') ]",
			"//a[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'export') ]",
			"//a[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'exportar') ]",
		]

		while time.time() < end:
			for xp in xpaths:
				try:
					els = driver.find_elements(By.XPATH, xp)
				except Exception:
					els = []
				if not els:
					continue
				for el in els:
					try:
						if not el.is_displayed():
							continue
					except Exception:
						pass
					try:
						driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
					except Exception:
						pass
					try:
						driver.execute_script("arguments[0].click();", el)
						LOG.info("click_export_link_with_fallback: clicado anchor por XPath: %s", xp)
						return True
					except Exception:
						try:
							ActionChains(driver).move_to_element(el).click().perform()
							LOG.info("click_export_link_with_fallback: clicado anchor via ActionChains por XPath: %s", xp)
							return True
						except Exception:
							LOG.debug("click_export_link_with_fallback: fallo al clicar anchor por XPath: %s", xp, exc_info=True)
			time.sleep(0.3)

		LOG.debug('click_export_link_with_fallback: no se encontró anchor de descarga dentro del timeout')
		return False
	except Exception:
		LOG.exception('click_export_link_with_fallback: excepción inesperada')
		return False


def bring_browser_to_front(driver: webdriver.Chrome, url_substring: str | None = None) -> bool:
	try:
		try:
			driver.execute_script("window.focus();")
		except Exception:
			pass

		try:
			from pywinauto import Desktop
		except Exception:
			return False

		try:
			desktop = Desktop(backend="uia")
			windows = desktop.windows()
			target = None
			us = url_substring.lower() if url_substring else None
			for w in windows:
				try:
					title = (w.window_text() or '').lower()
					if not title:
						continue
					if us and us in title:
						target = w
						break
					if not us and ('chrome' in title or 'qlik' in title):
						target = w
						break
				except Exception:
					continue

			if target:
				try:
					target.set_focus()
					return True
				except Exception:
					return False
		except Exception:
			return False

		return False
	except Exception:
		return False


def find_latest_downloaded_file(directory: str, pattern: str = '*.xlsx', since_ts: float | None = None, timeout: float = 30.0) -> str | None:
	end = time.time() + float(timeout)
	d = Path(directory).expanduser()
	if not d.exists():
		return None

	while time.time() < end:
		files = list(d.glob(pattern))
		if not files:
			time.sleep(0.5)
			continue
		files_sorted = sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)
		candidate = files_sorted[0]
		try:
			mtime = candidate.stat().st_mtime
		except Exception:
			time.sleep(0.5)
			continue
		if since_ts is None or mtime >= (since_ts - 2.0):
			return str(candidate)
		time.sleep(0.5)
	return None


def extract_excel_contents(path: str) -> dict | None:
	try:
		try:
			from openpyxl import load_workbook
			wb = load_workbook(path, data_only=True)
			out = {}
			for sheet in wb.sheetnames:
				ws = wb[sheet]
				rows = list(ws.rows)
				if not rows:
					out[sheet] = []
					continue
				headers = [(c.value if c.value is not None else f'col{i}') for i, c in enumerate(rows[0], start=1)]
				data = []
				for r in rows[1:]:
					rowd = {}
					for h, cell in zip(headers, r):
						rowd[str(h)] = str(cell.value) if cell.value is not None else ''
					data.append(rowd)
				out[sheet] = data
			return out
		except Exception:
			try:
				import pandas as pd
				xls = pd.ExcelFile(path)
				result = {}
				for sheet in xls.sheet_names:
					df = pd.read_excel(xls, sheet_name=sheet, dtype=str)
					df = df.fillna('')
					result[sheet] = df.to_dict(orient='records')
				return result
			except Exception:
				return None
	except Exception:
		return None


def upload_to_google_sheets(extracted: dict, spreadsheet_id: str, credentials_json_path: str, clear: bool = True, target_sheet: str | None = 'Sheet2') -> bool:
	try:
		try:
			import gspread
			from google.oauth2 import service_account
		except Exception:
			return False

		scopes = [
			'https://www.googleapis.com/auth/spreadsheets',
			'https://www.googleapis.com/auth/drive'
		]
		creds = service_account.Credentials.from_service_account_file(credentials_json_path, scopes=scopes)
		client = gspread.authorize(creds)
		sh = client.open_by_key(spreadsheet_id)

		if target_sheet:
			try:
				first_sheet = next(iter(extracted.keys()))
				rows = extracted.get(first_sheet, [])
				safe_name = str(target_sheet)[:100]
				try:
					ws = sh.worksheet(safe_name)
				except Exception:
					ws = sh.add_worksheet(title=safe_name, rows=max(100, len(rows) + 5), cols=20)

				if clear:
					try:
						rc = getattr(ws, 'row_count', None)
						if rc and isinstance(rc, int) and rc > 1:
							ws.batch_clear([f"2:{rc}"])
						else:
							ws.batch_clear(["2:1000"])
					except Exception:
						pass

				date_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
				
				if rows:
					headers = list(rows[0].keys()) if rows else []
					table_data = []
					for r in rows:
						row = [date_str]
						for h in headers:
							row.append(r.get(h, ''))
						table_data.append(row)
					
					try:
						ws.update('A2', table_data)
						LOG.info('upload_to_google_sheets: hoja %s actualizada (%d filas)', safe_name, len(rows))
					except Exception:
						pass
			except Exception:
				pass
		
		return True
	except Exception:
		return False


# ========== AQUÍ PEGAS LA NUEVA FUNCIÓN ==========

def detect_data_type_and_upload(extracted: dict, spreadsheet_id: str, credentials_json_path: str, clear: bool = True) -> bool:
	"""
	Detecta el tipo de datos (PDV, Zona o Clasificacion) y distribuye en hojas correspondientes
	
	Si el archivo contiene:
	- Columna A: "PDV" y Columna B: "Nivel de Servicio" → Sube a "Puntos_Venta"
	- Columna A: "Zona" y Columna B: "Nivel de Servicio" → Sube a "Zonas"
	- Columna A: "Clasificacion" y Columna B: "Nivel de Servicio" → Sube a "Clasificacion"
	"""
	try:
		try:
			import gspread
			from google.oauth2 import service_account
		except Exception:
			LOG.error("Falta gspread o google-auth")
			return False

		scopes = [
			'https://www.googleapis.com/auth/spreadsheets',
			'https://www.googleapis.com/auth/drive'
		]
		creds = service_account.Credentials.from_service_account_file(credentials_json_path, scopes=scopes)
		client = gspread.authorize(creds)
		sh = client.open_by_key(spreadsheet_id)

		# ====== PROCESAMIENTO DE CADA SHEET EXTRAÍDO ======
		for sheet_name, rows in extracted.items():
			if not rows:
				LOG.info(f"Sheet {sheet_name} está vacío, saltando")
				continue

			# Obtener headers (primera fila)
			headers = list(rows[0].keys()) if rows else []
			
			if len(headers) < 2:
				LOG.warning(f"Sheet {sheet_name} tiene menos de 2 columnas, saltando")
				continue

			# ====== DETECTAR TIPO DE DATOS ======
			header_col_a = str(headers[0]).strip().upper()
			header_col_b = str(headers[1]).strip().upper()

			LOG.info(f"Detectados headers: Col A='{header_col_a}', Col B='{header_col_b}'")

			# Determine target sheet and column mapping
			target_sheet = None
			is_pdv = False
			is_zona = False
			is_clasificacion = False

			# Condición 1: PDV (Puntos de Venta)
			if "PDV" in header_col_a and "NIVEL" in header_col_b and "SERVICIO" in header_col_b:
				target_sheet = "Puntos_Venta"
				is_pdv = True
				LOG.info(f"✓ Detectado: PUNTOS DE VENTA (PDV) → Hoja: {target_sheet}")

			# Condición 2: ZONA
			elif "ZONA" in header_col_a and "NIVEL" in header_col_b and "SERVICIO" in header_col_b:
				target_sheet = "Zonas"
				is_zona = True
				LOG.info(f"✓ Detectado: ZONAS → Hoja: {target_sheet}")

			# Condición 3: CLASIFICACION
			elif "CLASIFICACION" in header_col_a and "NIVEL" in header_col_b and "SERVICIO" in header_col_b:
				target_sheet = "Clasificacion"
				is_clasificacion = True
				LOG.info(f"✓ Detectado: CLASIFICACIÓN → Hoja: {target_sheet}")

			else:
				LOG.warning(f"No se pudo identificar el tipo de datos. Headers: {headers}")
				continue

			# ====== PROCESAR Y SUBIR A HOJA DESTINO ======
			if target_sheet:
				try:
					# Obtener o crear worksheet
					try:
						ws = sh.worksheet(target_sheet)
						LOG.info(f"Worksheet '{target_sheet}' encontrada")
					except Exception:
						ws = sh.add_worksheet(title=target_sheet, rows=max(100, len(rows) + 5), cols=20)
						LOG.info(f"Worksheet '{target_sheet}' creada")

					# Limpiar filas previas si es necesario
					if clear:
						try:
							rc = getattr(ws, 'row_count', None)
							if rc and isinstance(rc, int) and rc > 1:
								ws.batch_clear([f"2:{rc}"])
							else:
								ws.batch_clear(["2:1000"])
							LOG.info(f"Worksheet '{target_sheet}' limpiada")
						except Exception as e:
							LOG.warning(f"Error al limpiar worksheet: {e}")

					# ====== PREPARAR DATOS PARA SUBIDA ======
					date_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
					table_data = []

					if is_pdv:
						# Para PDV: Fecha | PDV | Nivel de Servicio
						# Los datos van en columnas B y C
						for r in rows:
							valor_col_a = r.get(headers[0], '')  # PDV
							valor_col_b = r.get(headers[1], '')  # Nivel de Servicio
							row = [date_str, valor_col_a, valor_col_b]
							table_data.append(row)
						
						range_update = "A2"  # Desde A2 (fila 2)
						LOG.info(f"Subiendo {len(table_data)} filas a '{target_sheet}' (columnas B y C)")

					elif is_zona:
						# Para ZONA: Fecha | Zona | Nivel de Servicio
						# Los datos van en columnas B y C
						for r in rows:
							valor_col_a = r.get(headers[0], '')  # Zona
							valor_col_b = r.get(headers[1], '')  # Nivel de Servicio
							row = [date_str, valor_col_a, valor_col_b]
							table_data.append(row)
						
						range_update = "A2"  # Desde A2 (fila 2)
						LOG.info(f"Subiendo {len(table_data)} filas a '{target_sheet}' (columnas B y C)")

					elif is_clasificacion:
						# Para CLASIFICACION: Fecha | Clasificacion | Nivel de Servicio
						# Los datos van en columnas B y C
						for r in rows:
							valor_col_a = r.get(headers[0], '')  # Clasificacion
							valor_col_b = r.get(headers[1], '')  # Nivel de Servicio
							row = [date_str, valor_col_a, valor_col_b]
							table_data.append(row)
						
						range_update = "A2"  # Desde A2 (fila 2)
						LOG.info(f"Subiendo {len(table_data)} filas a '{target_sheet}' (columnas B y C)")

					# ====== EJECUTAR SUBIDA ======
					if table_data:
						try:
							ws.update(range_update, table_data)
							LOG.info(f'✓ Sheet "{target_sheet}" actualizada con {len(table_data)} filas')
						except Exception as e:
							LOG.error(f"Error al actualizar sheet: {e}", exc_info=True)
							return False
					else:
						LOG.warning(f"No hay datos para subir a '{target_sheet}'")

				except Exception as e:
					LOG.error(f"Error procesando hoja {target_sheet}: {e}", exc_info=True)
					continue

		return True

	except Exception as e:
		LOG.error(f"Error en detect_data_type_and_upload: {e}", exc_info=True)
		return False


def mes_numero_a_abreviatura(mes_numero: int) -> str:
	"""Convierte número de mes (1-12) a abreviatura (Ene, Feb, Mar...)"""
	meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
	return meses[mes_numero - 1] if 1 <= mes_numero <= 12 else 'Ene'


def seleccionar_mes_o_ano(driver: webdriver.Chrome, selector_contenedor: str, valor_a_escribir: str) -> bool:
	"""
	Hace click en un contenedor, escribe un valor (mes o año),
	envía 2 TAB + SPACE y exporta.
	"""
	try:
		# Click en el contenedor
		if not click_button_by_selector(driver, selector_contenedor, timeout=8.0):
			LOG.info("No se pudo clickear el contenedor: %s", selector_contenedor)
			return False
		time.sleep(4)  # Esperar carga

		# Escribir valor (mes abreviado o año)
		actions = ActionChains(driver)
		actions.send_keys(valor_a_escribir).perform()
		time.sleep(1.5)

		# Dos TAB y un SPACE
		actions.send_keys(Keys.TAB).send_keys(Keys.TAB).send_keys(Keys.SPACE).perform()
		time.sleep(2)

		LOG.info("Seleccionado: %s", valor_a_escribir)
		return True
	except Exception:
		LOG.exception("Error en seleccionar_mes_o_ano")
		return False


def click_boton_exportar_toolbar(driver: webdriver.Chrome) -> bool:
	"""
	Click en el botón exportar del toolbar.
	"""
	try:
		exportar_selector = (
			"#actions-toolbar > div.njs-b447-Grid-root.njs-b447-Grid-container."
			"njs-b447-Grid-item.njs-b447-Grid-wrap-xs-nowrap.actions-toolbar-default-actions."
			"css-3cuy5k > div:nth-child(3) > button > i > svg"
		)
		if click_button_by_selector(driver, exportar_selector, timeout=8.0):
			LOG.info("Botón exportar clickeado")
			time.sleep(2)
			return True
		LOG.warning("No se pudo clickear botón exportar")
		return False
	except Exception:
		LOG.exception("Error en click_boton_exportar_toolbar")
		return False
# ========== FIN DE LA NUEVA FUNCIÓN ==========






def grid_listo(driver: webdriver.Chrome, selector: str, selector_type: str = 'CSS_SELECTOR', timeout: float = 20.0) -> bool:
	try:
		end = time.time() + float(timeout)
		
		if selector_type.upper() == 'XPATH':
			by = By.XPATH
		else:
			by = By.CSS_SELECTOR
		
		while time.time() < end:
			try:
				el = driver.find_element(by, selector)
				if el.is_displayed():
					return True
			except Exception:
				pass
			time.sleep(0.5)
		return False
	except Exception:
		return False


def run_once() -> None:
	# URLs ORIGINALES DE qliktabsLVL
	url = (
		"https://qlik.copservir.com/sense/app/57d34c58-0147-49a3-ba25-5ab93cd53935/sheet/1b65938b-8bbb-4f9a-80b7-da596cf645d9/state/analysis?qlikTicket=yp_QvgbP53N6A11R"
	)
	
	# CREDENCIALES ORIGINALES
	username = "Qlikzona29"
	password = "pF2A3f2x*"

	# GOOGLE SHEETS ORIGINALES
	default_sa = r'C:\Users\cmarroquin\Music\qliktabs-LVL servic\prueba-de-gmail-486215-55c5b9a4012b.json'
	default_sid = '1w64tHFJuoLqsoxTwotfCJ-fOVEJOO3oKZiIjaIflBNU'
	sa = _os.environ.get('GOOGLE_SERVICE_ACCOUNT_JSON', default_sa)
	sid = _os.environ.get('GOOGLE_SHEET_ID', default_sid)
	target_hoja1 = _os.environ.get('GOOGLE_SHEET_TAB', 'Hoja ')
	target_hoja2 = _os.environ.get('GOOGLE_SHEET_TAB_2', 'Hoja 2')

	logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
	LOG.info('Starting Qlik autofill (single run)')
	driver = setup_driver()
	submit_sent = False
	wrote_pwd = False
	try:
		LOG.info("Opening %s", url)
		driver.get(url)
		time.sleep(25)
		
		try:
			bring_browser_to_front(driver)
		except Exception:
			pass

		try:
			WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#grid > div:nth-child(9)")))
			focus_on_selector(driver, "#grid > div:nth-child(9)", timeout=3.0)
		except Exception:
			time.sleep(2)

		try:
			try:
				bring_browser_to_front(driver)
			except Exception:
				pass
			
			type_like_keyboard(driver, username, delay=0.08, click_first=True)

			success = False
			try:
				active = driver.switch_to.active_element
				val = active.get_attribute('value') or ''
				if username in val:
					success = True
			except Exception:
				pass

			try:
				parsed = urllib.parse.urlparse(url)
				host = parsed.netloc
			except Exception:
				host = 'qlik'

			if not success:
				send_text_via_system(driver, username, delay=0.08)
				try:
					send_keys_via_pywinauto('{TAB}', host)
				except Exception:
					pass

			if send_text_via_system(driver, password, delay=0.08):
				wrote_pwd = True
				try:
					if send_keys_via_pywinauto('{TAB}{ENTER}', host):
						submit_sent = True
					else:
						pass
				except Exception:
					pass
		except Exception:
			pass

		try:
			_keep_open_seconds = 30
			if submit_sent or wrote_pwd:
				time.sleep(_keep_open_seconds)
	
	# ====== FILTRAR POR MES Y AÑO ACTUAL ======
			try:
				hoy = datetime.now()
				mes_actual = hoy.month
				ano_actual = str(hoy.year)
				
				mes_abrev = mes_numero_a_abreviatura(mes_actual)
				
				# ========== SELECTOR MES ==========
				selector_mes = (
					"#bc155f64-af38-47fe-8ffe-9e5093c2852e_content > div > div > "
					"div:nth-child(2) > div > div > div > div > div > "
					"div.MuiGrid-root.MuiGrid-container.css-q0qbej"
				)
				
				if seleccionar_mes_o_ano(driver, selector_mes, mes_abrev):
					click_boton_exportar_toolbar(driver)
					time.sleep(3)
				
				# ========== SELECTOR AÑO ==========
				selector_ano = (
					"#bc155f64-af38-47fe-8ffe-9e5093c2852e_content > div > div > "
					"div:nth-child(1) > div > div > div > div > div > "
					"div.MuiGrid-root.MuiGrid-container.css-q0qbej"
				)
				
				if seleccionar_mes_o_ano(driver, selector_ano, ano_actual):
					click_boton_exportar_toolbar(driver)
					time.sleep(3)

				LOG.info("Filtros de mes (%s) y año (%s) aplicados", mes_abrev, ano_actual)
				time.sleep(3)  # Esperar a que carguen los datos filtrados

			except Exception as e:
				LOG.error(f"Error aplicando filtros: {e}", exc_info=True)

			# ===== GRID 0 (CLASIFICACION) - PRIMERO =====
			try:
				grid_sel_0 = "#grid > div:nth-child(8)"
				WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.CSS_SELECTOR, grid_sel_0)))
				
				if not grid_listo(driver, grid_sel_0, timeout=20):
					LOG.info("Grid 0 (Clasificacion) no está listo, saltando")
				else:
					try:
						bring_browser_to_front(driver)
					except Exception:
						pass
					
					if hover_on_selector(driver, grid_sel_0, timeout=5.0):
						time.sleep(15)
						try:
							btn_sel_0 = (
								'#grid > div:nth-child(8) > '
								'div.object-and-panel-wrapper > div > '
								'div.ng-isolate-scope.detached-object-nav-wrapper > div '
								'button[tid="nav-menu-move"]'
							)
							if click_button_by_selector(driver, btn_sel_0, timeout=5.0):
								time.sleep(2)
								try:
									if click_button_by_selector(driver, '#export-group', timeout=5.0):
										time.sleep(0.6)
										if click_button_by_selector(driver, '#export', timeout=5.0):
											time.sleep(0.6)
											try:
												download_start_ts_0 = time.time()
												if click_export_url(driver, selector='a.export-url', timeout=10.0):
													time.sleep(8)

													downloads_dir_0 = os.path.join(Path.home(), 'Downloads')
													found_0 = find_latest_downloaded_file(downloads_dir_0, pattern='*.xlsx', since_ts=download_start_ts_0, timeout=30.0)
													if found_0:
														LOG.info('Archivo Grid 0 (Clasificacion): %s', found_0)
														extracted_0 = extract_excel_contents(found_0)
														if extracted_0 is not None:
															out_file_0 = Path('exported_data_grid0.json')
															try:
																with out_file_0.open('w', encoding='utf-8') as fh:
																	json.dump(extracted_0, fh, ensure_ascii=False, indent=2)
															except Exception:
																pass
															try:
																detect_data_type_and_upload(extracted_0, sid, sa, clear=True)
															except Exception as e:
																LOG.error(f"Error en detección de datos Grid 0: {e}", exc_info=True)
															try:
																p_0 = Path(found_0)
																if p_0.exists():
																	p_0.unlink()
															except Exception:
																pass
														
														# Cerrar diálogo
														try:
															LOG.info("Clickeando botón de diálogo Grid 0")
															if click_button_by_selector(driver, "#export-dialog > div > div.lui-dialog__footer > button", timeout=5.0):
																LOG.info("Botón diálogo Grid 0 clicado")
																time.sleep(2)
															else:
																LOG.info("No se pudo clickear botón diálogo Grid 0")
														except Exception:
															LOG.debug('Error al clickear dialog Grid 0', exc_info=True)
											except Exception:
												pass
								except Exception:
									pass
						except Exception:
							pass
			except Exception:
				LOG.debug("Error en Grid 0", exc_info=True)

			# ===== GRID 1 =====
			try:
				grid_sel = "#grid > div:nth-child(9)"
				WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.CSS_SELECTOR, grid_sel)))
				
				if not grid_listo(driver, grid_sel, timeout=20):
					LOG.info("Grid 1 no está listo, saltando")
				else:
					try:
						bring_browser_to_front(driver)
					except Exception:
						pass
					
					if hover_on_selector(driver, grid_sel, timeout=5.0):
						time.sleep(15)
						try:
							btn_sel = (
								'#grid > div:nth-child(9) > '
								'div.object-and-panel-wrapper > div > '
								'div.ng-isolate-scope.detached-object-nav-wrapper > div '
								'button[tid="nav-menu-move"]'
							)
							if click_button_by_selector(driver, btn_sel, timeout=5.0):
								time.sleep(2)
								try:
									if click_button_by_selector(driver, '#export-group', timeout=5.0):
										time.sleep(0.6)
										if click_button_by_selector(driver, '#export', timeout=5.0):
											time.sleep(0.6)
											try:
												download_start_ts = time.time()
												if click_export_url(driver, selector='a.export-url', timeout=10.0):
													time.sleep(8)

													downloads_dir = os.path.join(Path.home(), 'Downloads')
													found = find_latest_downloaded_file(downloads_dir, pattern='*.xlsx', since_ts=download_start_ts, timeout=30.0)
													if found:
														LOG.info('Archivo Grid 1: %s', found)
														extracted = extract_excel_contents(found)
														if extracted is not None:
															out_file = Path('exported_data_grid1.json')
															try:
																with out_file.open('w', encoding='utf-8') as fh:
																	json.dump(extracted, fh, ensure_ascii=False, indent=2)
															except Exception:
																pass
															try:
																detect_data_type_and_upload(extracted, sid, sa, clear=True)
															except Exception as e:
																LOG.error(f"Error en detección de datos Grid 1: {e}", exc_info=True)
															try:
																p = Path(found)
																if p.exists():
																	p.unlink()
															except Exception:
																pass
														
														# Cerrar diálogo
														try:
															LOG.info("Clickeando botón de diálogo Grid 1")
															if click_button_by_selector(driver, "#export-dialog > div > div.lui-dialog__footer > button", timeout=5.0):
																LOG.info("Botón diálogo Grid 1 clicado")
																time.sleep(2)
															else:
																LOG.info("No se pudo clickear botón diálogo Grid 1")
														except Exception:
															LOG.debug('Error al clickear dialog Grid 1', exc_info=True)
											except Exception:
												pass
								except Exception:
									pass
						except Exception:
							pass
			except Exception:
				LOG.debug("Error en Grid 1", exc_info=True)

			# ===== GRID 2 =====
			try:
				grid_sel_2 = "#grid > div:nth-child(11)"
				WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.CSS_SELECTOR, grid_sel_2)))
				
				if not grid_listo(driver, grid_sel_2, timeout=20):
					LOG.info("Grid 2 no está listo, saltando")
				else:
					try:
						bring_browser_to_front(driver)
					except Exception:
						pass
					
					if hover_on_selector(driver, grid_sel_2, timeout=5.0):
						time.sleep(15)
						try:
							btn_sel_2 = (
								'#grid > div:nth-child(11) > '
								'div.object-and-panel-wrapper > div > '
								'div.ng-isolate-scope.detached-object-nav-wrapper > div '
								'button[tid="nav-menu-move"]'
							)
							if click_button_by_selector(driver, btn_sel_2, timeout=5.0):
								time.sleep(2)
								try:
									if click_button_by_selector(driver, '#export-group', timeout=5.0):
										time.sleep(0.6)
										if click_button_by_selector(driver, '#export', timeout=5.0):
											time.sleep(0.6)
											try:
												download_start_ts_2 = time.time()
												if click_export_url(driver, selector='a.export-url', timeout=10.0):
													time.sleep(8)

													downloads_dir_2 = os.path.join(Path.home(), 'Downloads')
													found_2 = find_latest_downloaded_file(downloads_dir_2, pattern='*.xlsx', since_ts=download_start_ts_2, timeout=30.0)
													if found_2:
														LOG.info('Archivo Grid 2: %s', found_2)
														extracted_2 = extract_excel_contents(found_2)
														if extracted_2 is not None:
															out_file_2 = Path('exported_data_grid2.json')
															try:
																with out_file_2.open('w', encoding='utf-8') as fh:
																	json.dump(extracted_2, fh, ensure_ascii=False, indent=2)
															except Exception:
																pass
															try:
																detect_data_type_and_upload(extracted_2, sid, sa, clear=True)
															except Exception as e:
																LOG.error(f"Error en detección de datos Grid 2: {e}", exc_info=True)
															try:
																p_2 = Path(found_2)
																if p_2.exists():
																	p_2.unlink()
															except Exception:
																pass
														
														# Cerrar diálogo
														try:
															LOG.info("Clickeando botón de diálogo Grid 2")
															if click_button_by_selector(driver, "#export-dialog > div > div.lui-dialog__footer > button", timeout=5.0):
																LOG.info("Botón diálogo Grid 2 clicado")
																time.sleep(2)
															else:
																LOG.info("No se pudo clickear botón diálogo Grid 2")
														except Exception:
															LOG.debug('Error al clickear dialog Grid 2', exc_info=True)
											except Exception:
												pass
								except Exception:
									pass
						except Exception:
							pass
			except Exception:
				LOG.debug("Error en Grid 2", exc_info=True)
		except Exception:
			pass
	finally:
		driver.quit()

def main() -> None:
	try:
		while True:
			try:
				run_once()
			except Exception:
				LOG.exception('run_once: excepción')

			now = datetime.now()
			next_run_6_exact = now.replace(hour=6, minute=0, second=0, microsecond=0)
			next_run_6 = now.replace(hour=6, minute=30, second=0, microsecond=0)
			next_run_12 = now.replace(hour=12, minute=30, second=0, microsecond=0)
			
			candidates = [next_run_6_exact, next_run_6]
			if now.weekday() in [5, 6]:
				candidates.append(next_run_12)
			future_candidates = [t for t in candidates if t > now]
			if future_candidates:
				next_run = min(future_candidates)
			else:
				next_run = min(candidates) + timedelta(days=1)
			
			wait_seconds = (next_run - now).total_seconds()
			LOG.info('Siguiente ejecución: %s', next_run.isoformat())

			try:
				time.sleep(wait_seconds)
			except KeyboardInterrupt:
				break
	except KeyboardInterrupt:
		LOG.info('Saliendo')


if __name__ == '__main__':
	main()